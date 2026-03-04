import openpyxl
import sqlite3
import pandas as pd
from typing import List
from src.stock_analyzer import StockAnalyzer

class SheetWorker:
    """
    A class for reading and updating Excel files with stock data.
    """

    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filename)
        self.symbols: List[str] = []

    def read_symbols(self) -> None:
        """Read symbols from the first sheet, column A, starting from row 2."""
        sheet = self.workbook.active
        self.symbols = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet[f'A{row}'].value
            if cell_value:
                self.symbols.append(cell_value)

    def get_symbols(self) -> List[str]:
        """Return the list of symbols."""
        return self.symbols

    def update_excel_from_db(self, analyst: StockAnalyzer) -> None:
        """Update the Excel file with data from the StockAnalyzer."""
        # Get summary data from analyst
        summary_df = analyst.read_from_sqlite('stocks.db')
        if summary_df is None:
            raise ValueError("No data available from database")

        if 'Data' not in self.workbook.sheetnames:
            self.workbook.create_sheet('Data')
        data_sheet = self.workbook['Data']

        # Clear existing data
        data_sheet.delete_rows(1, data_sheet.max_row)

        # Write DataFrame to sheet
        for r, (idx, row) in enumerate(summary_df.iterrows(), start=1):
            for c, (col_name, value) in enumerate(row.items(), start=1):
                if r == 1:
                    data_sheet.cell(row=r, column=c, value=col_name)
                data_sheet.cell(row=r+1, column=c, value=value)

        self.workbook.save(self.filename)
