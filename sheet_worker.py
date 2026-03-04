import yfinance as yf
from openpyxl import load_workbook
import logging
import pandas as pd
import pdb
class SheetWorker:
	def __init__(self,fn):
		self.wb = load_workbook(fn)
		self.fn = fn
		self.sheet = self.wb.active
		self.symbols = []
		self.prices = dict()
		self.targets = dict()

	def read_symbols(self):
		for cell in self.sheet['A']:
			if cell.value and cell.value not in ['name' or '$NAME']:
				self.symbols.append(cell.value)

	def get_symbols(self):
		return self.symbols

	def set_symbols(self, symbols):
		self.symbols = symbols

	def get_data(self):
		try:
			data = yf.Tickers(self.symbols)
			for symbol in self.symbols:
				try:
					ticker = data.tickers[symbol]
					self.prices[symbol] = ticker.fast_info['last_price']
					info = ticker.info
					self.targets[symbol] = {
					'mean': info.get('targetMeanPrice'),
					'high': info.get('targetHighPrice'),
					'low': info.get('targetLowPrice')
					}
				except Exception as e:
					logging.error(f"Error getting data for {symbol}: {str(e)}")
		except Exception as e:
			logging.error(f"Error initializing yfinance Tickers: {str(e)}")

	def get_info(self):
		"""Fetches stock data and updates the Excel sheet with prices and targets."""
		for cell in self.sheet['A']:
			if cell.value in self.prices:
				#TODO make col dynamic
				price_cell =  self.sheet.cell(row=cell.row, column=2)
	#			print(cell.value)
				price_cell.value = self.prices[cell.value]
			if cell.value in self.targets:
				target_cell = self.sheet.cell(row=cell.row, column=3)
				target_cell.value = self.targets[cell.value].get('mean')

	def save(self,fn=None):
		if not fn:
			fn=self.fn
		self.wb.save(fn)

	def update_excel_from_db(self, analyst):
		wb = load_workbook(self.fn)
		ws = wb['Summary']  # Assume a sheet for summaries
		
		# Clear existing data if needed
		for row in ws.iter_rows(min_row=2):
			for cell in row:
				cell.value = None
		
		row_num = 2
		for symbol, data in analyst.data.items():
			latest = data.iloc[-1]
			fund = analyst.fundamental_data.get(symbol, {})
			ws.cell(row=row_num, column=1).value = symbol
			ws.cell(row=row_num, column=2).value = latest['Close']
			ws.cell(row=row_num, column=3).value = latest['RSI']
			row_num += 1

		ws_main = wb['Sheet1']
		symbol_to_row = {ws_main.cell(row=i, column=1).value: i for i in range(2, ws_main.max_row + 1)}
		for symbol in symbol_to_row:
			if symbol in analyst.data and not analyst.data[symbol].empty:
				row = symbol_to_row[symbol]
				latest = analyst.data[symbol].iloc[-1]
				ws_main.cell(row=row, column=2).value = latest['Close']  # last close
				ws_main.cell(row=row, column=3).value = analyst.targets.get(symbol, {}).get('mean')  # mean target
				ws_main.cell(row=row, column=26).value = analyst.targets.get(symbol, {}).get('mean')  # mean target
			else:
				logging.warning(f"Skipping Sheet1 update for {symbol}: No data available")

		wb.save(self.fn)
		